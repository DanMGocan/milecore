import io
import os
import re

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from backend.auth import InstanceContext, get_current_instance
from backend.database import (
    execute_query,
    get_all_table_rows,
    get_table_rows,
    get_table_schema,
    get_tables,
)

router = APIRouter(prefix="/tables")

SKIP_TABLES = {"audit_log", "app_settings", "chat_sessions", "chat_messages", "pending_approvals"}


@router.get("/export")
async def export_excel(ctx: InstanceContext = Depends(get_current_instance)):
    wb = Workbook()
    wb.remove(wb.active)
    for table in get_tables(instance_id=ctx.instance_id):
        if table in SKIP_TABLES:
            continue
        columns, rows = get_all_table_rows(table, instance_id=ctx.instance_id)
        ws = wb.create_sheet(title=table)
        ws.append(columns)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=truecore_export.xlsx"},
    )


@router.post("/import")
async def import_excel(file: UploadFile = File(...), ctx: InstanceContext = Depends(get_current_instance)):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    existing_tables = set(get_tables(instance_id=ctx.instance_id))
    summary = {"tables_imported": 0, "rows_per_table": {}, "skipped_sheets": [], "errors": []}

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_TABLES or sheet_name not in existing_tables:
                summary["skipped_sheets"].append(sheet_name)
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                summary["skipped_sheets"].append(sheet_name)
                continue
            headers = rows[0]
            if not all(isinstance(h, str) and re.match(r"^\w+$", h) for h in headers):
                summary["errors"].append(f"{sheet_name}: invalid column names")
                continue
            data_rows = rows[1:]
            # Clear existing data for this instance
            result = execute_query(
                f"DELETE FROM {sheet_name} WHERE instance_id = ?",
                [ctx.instance_id],
                instance_id=ctx.instance_id,
            )
            if "error" in result:
                summary["errors"].append(f"{sheet_name}: {result['error']}")
                continue
            # Insert rows with instance_id
            placeholders = ", ".join(["?"] * len(headers)) + ", ?"
            col_names = ", ".join(headers) + ", instance_id"
            inserted = 0
            for row in data_rows:
                values = list(row) + [ctx.instance_id]
                result = execute_query(
                    f"INSERT INTO {sheet_name} ({col_names}) VALUES ({placeholders})",
                    values,
                    instance_id=ctx.instance_id,
                )
                if "error" in result:
                    summary["errors"].append(f"{sheet_name} row: {result['error']}")
                else:
                    inserted += 1
            summary["rows_per_table"][sheet_name] = inserted
            summary["tables_imported"] += 1
    finally:
        wb.close()
    return summary


@router.post("/import-merge")
async def import_merge(file: UploadFile = File(...), ctx: InstanceContext = Depends(get_current_instance)):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    contents = await file.read()
    existing_tables = set(get_tables(instance_id=ctx.instance_id))
    summary = {"tables_imported": 0, "rows_per_table": {}, "skipped_sheets": [], "errors": []}

    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_TABLES or sheet_name not in existing_tables:
                summary["skipped_sheets"].append(sheet_name)
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                summary["skipped_sheets"].append(sheet_name)
                continue
            headers = rows[0]
            if not all(isinstance(h, str) and re.match(r"^\w+$", h) for h in headers):
                summary["errors"].append(f"{sheet_name}: invalid column names")
                continue
            data_rows = rows[1:]
            placeholders = ", ".join(["?"] * len(headers)) + ", ?"
            col_names = ", ".join(headers) + ", instance_id"
            inserted = 0
            for row in data_rows:
                values = list(row) + [ctx.instance_id]
                result = execute_query(
                    f"INSERT INTO {sheet_name} ({col_names}) VALUES ({placeholders}) ON CONFLICT DO NOTHING",
                    values,
                    instance_id=ctx.instance_id,
                )
                if "error" in result:
                    summary["errors"].append(f"{sheet_name} row: {result['error']}")
                else:
                    inserted += result.get("rowcount", 0)
            summary["rows_per_table"][sheet_name] = inserted
            summary["tables_imported"] += 1
    finally:
        wb.close()

    return summary


def _validate_name(name: str) -> str:
    if not re.match(r"^\w+$", name):
        raise HTTPException(status_code=400, detail="Invalid name")
    return name


# --- Table operations ---


class CreateTableRequest(BaseModel):
    name: str
    columns: list[dict]  # [{"name": "col1", "type": "TEXT", "notnull": False, "default": None}]


class AddColumnRequest(BaseModel):
    name: str
    type: str = "TEXT"
    default: str | None = None


class InsertRowRequest(BaseModel):
    data: dict


@router.get("")
async def list_tables(ctx: InstanceContext = Depends(get_current_instance)):
    return {"tables": get_tables(instance_id=ctx.instance_id)}


@router.post("")
async def create_table(req: CreateTableRequest, ctx: InstanceContext = Depends(get_current_instance)):
    name = _validate_name(req.name)
    if not req.columns:
        raise HTTPException(status_code=400, detail="At least one column is required")

    col_defs = []
    for col in req.columns:
        col_name = _validate_name(col["name"])
        col_type = col.get("type", "TEXT")
        parts = [col_name, col_type]
        if col.get("notnull"):
            parts.append("NOT NULL")
        if col.get("default") is not None:
            parts.append(f"DEFAULT {col['default']!r}")
        col_defs.append(" ".join(parts))

    col_defs.insert(0, "id BIGSERIAL PRIMARY KEY")
    col_defs.append("instance_id BIGINT NOT NULL REFERENCES instances(id)")
    sql = f"CREATE TABLE {name} ({', '.join(col_defs)})"
    result = execute_query(sql, instance_id=ctx.instance_id)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return {"message": f"Table '{name}' created", "sql": sql}


@router.delete("/{table_name}")
async def drop_table(table_name: str, ctx: InstanceContext = Depends(get_current_instance)):
    name = _validate_name(table_name)
    result = execute_query(f"DROP TABLE IF EXISTS {name}", instance_id=ctx.instance_id)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return {"message": f"Table '{name}' dropped"}


# --- Schema operations ---


@router.get("/{table_name}/schema")
async def table_schema(table_name: str, ctx: InstanceContext = Depends(get_current_instance)):
    name = _validate_name(table_name)
    schema = get_table_schema(name)
    if not schema:
        raise HTTPException(status_code=404, detail="Table not found")
    return {"table": name, "columns": schema}


@router.post("/{table_name}/columns")
async def add_column(table_name: str, req: AddColumnRequest, ctx: InstanceContext = Depends(get_current_instance)):
    table = _validate_name(table_name)
    col = _validate_name(req.name)
    sql = f"ALTER TABLE {table} ADD COLUMN {col} {req.type}"
    if req.default is not None:
        sql += f" DEFAULT {req.default!r}"
    result = execute_query(sql, instance_id=ctx.instance_id)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return {"message": f"Column '{col}' added to '{table}'"}


# --- Row operations ---


@router.get("/{table_name}/rows")
async def table_rows(
    table_name: str,
    limit: int = Query(default=50, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    ctx: InstanceContext = Depends(get_current_instance),
):
    name = _validate_name(table_name)
    result = get_table_rows(name, limit, offset, instance_id=ctx.instance_id)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.post("/{table_name}/rows")
async def insert_row(table_name: str, req: InsertRowRequest, ctx: InstanceContext = Depends(get_current_instance)):
    name = _validate_name(table_name)
    if not req.data:
        raise HTTPException(status_code=400, detail="No data provided")

    for key in req.data.keys():
        if not re.match(r"^\w+$", key):
            raise HTTPException(status_code=400, detail=f"Invalid column name: {key}")

    # Add instance_id to the data
    data = {**req.data, "instance_id": ctx.instance_id}
    columns = ", ".join(data.keys())
    placeholders = ", ".join(["?"] * len(data))
    values = list(data.values())

    sql = f"INSERT INTO {name} ({columns}) VALUES ({placeholders})"
    result = execute_query(sql, values, instance_id=ctx.instance_id)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return {"message": "Row inserted", "id": result.get("lastrowid")}


@router.delete("/{table_name}/rows/{row_id}")
async def delete_row(table_name: str, row_id: int, ctx: InstanceContext = Depends(get_current_instance)):
    name = _validate_name(table_name)
    result = execute_query(
        f"DELETE FROM {name} WHERE id = ? AND instance_id = ?",
        [row_id, ctx.instance_id],
        instance_id=ctx.instance_id,
    )
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    if result.get("rowcount", 0) == 0:
        raise HTTPException(status_code=404, detail="Row not found")
    return {"message": f"Row {row_id} deleted from '{name}'"}
