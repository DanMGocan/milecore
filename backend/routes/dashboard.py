import os
import re

from fastapi import APIRouter, Query
from pydantic import BaseModel

from backend.claude_client import clear_prompt_cache, clear_schema_cache
from backend.config import APP_URL, SCHEMA_PATH
from backend.email_sender import send_email
from backend.database import _lock, get_connection, reset_db
from initial_seed import seed_initial_data

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

router = APIRouter(prefix="/dashboard")


def _query(sql: str, params: list | None = None) -> list[dict]:
    conn = get_connection()
    with _lock:
        cursor = conn.execute(sql, params or [])
        columns = [d[0] for d in cursor.description] if cursor.description else []
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _count(sql: str) -> int:
    conn = get_connection()
    with _lock:
        return conn.execute(sql).fetchone()[0]


@router.get("/overview")
async def overview():
    row = _query(
        "SELECT "
        "(SELECT COUNT(*) FROM assets WHERE lifecycle_status = 'active') AS active_assets, "
        "(SELECT COUNT(*) FROM requests WHERE status IN ('open', 'in_progress')) AS open_requests, "
        "(SELECT COUNT(*) FROM technical_issues WHERE resolution IS NULL) AS open_issues, "
        "(SELECT COUNT(*) FROM events WHERE DATE(start_time) BETWEEN DATE('now', 'weekday 1', '-7 days') AND DATE('now', '+7 days')) AS events_this_week, "
        "(SELECT COUNT(*) FROM technical_issues WHERE important=1) + "
        "(SELECT COUNT(*) FROM requests WHERE important=1) + "
        "(SELECT COUNT(*) FROM events WHERE important=1) + "
        "(SELECT COUNT(*) FROM notes WHERE important=1) + "
        "(SELECT COUNT(*) FROM changes WHERE important=1) + "
        "(SELECT COUNT(*) FROM work_logs WHERE important=1) + "
        "(SELECT COUNT(*) FROM assets WHERE important=1) + "
        "(SELECT COUNT(*) FROM inventory_transactions WHERE important=1) AS important_items"
    )[0]
    push = _query("SELECT value FROM app_settings WHERE key = 'last_push_at'")
    row["last_push"] = push[0]["value"] if push else None
    return row


@router.get("/assets-by-period")
async def assets_by_period():
    data = _query(
        "SELECT DATE(created_at) as date, COUNT(*) as count FROM assets "
        "WHERE created_at IS NOT NULL "
        "GROUP BY DATE(created_at) ORDER BY date DESC LIMIT 30"
    )
    data.reverse()
    return {"data": data}


@router.get("/issues-summary")
async def issues_summary():
    by_status = _query(
        "SELECT status, COUNT(*) as count FROM requests GROUP BY status ORDER BY count DESC"
    )
    by_severity = _query(
        "SELECT severity, COUNT(*) as count FROM technical_issues "
        "WHERE severity IS NOT NULL GROUP BY severity ORDER BY count DESC"
    )
    return {"by_status": by_status, "by_severity": by_severity}


@router.get("/vendor-visits")
async def vendor_visits():
    data = _query(
        "SELECT status, COUNT(*) as count FROM events "
        "WHERE LOWER(event_type) LIKE '%vendor%' OR LOWER(event_type) LIKE '%visit%' "
        "GROUP BY status"
    )
    return {"data": data}


@router.get("/staff-per-site")
async def staff_per_site():
    data = _query(
        "SELECT s.name as site, COUNT(p.id) as count "
        "FROM sites s LEFT JOIN people p ON p.site_id = s.id AND p.employer_id IS NOT NULL AND p.status = 'active' "
        "GROUP BY s.id, s.name ORDER BY count DESC"
    )
    return {"data": data}


@router.post("/seed-demo")
async def seed_demo():
    """Seed the database with demo data from the xlsx file."""
    import openpyxl

    xlsx_path = os.path.join(_PROJECT_ROOT, "dummy_files", "demo_full_import.xlsx")
    if not os.path.exists(xlsx_path):
        return {"ok": False, "error": "Demo data file not found"}

    # Check if demo data has already been seeded (initial seed only has 1 company)
    conn = get_connection()
    row = conn.execute("SELECT COUNT(*) FROM companies").fetchone()
    if row and row[0] > 1:
        return {"ok": False, "already_seeded": True,
                "error": "The database already contains demo data. Please reset the database from the Dashboard page first."}

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        conn = get_connection()

        # Insert order respects FK constraints
        sheet_order = [
            "companies", "sites", "rooms", "people", "assets",
            "requests", "technical_issues", "events", "inventory_items",
        ]

        total_inserted = 0

        with _lock:
            for table_name in sheet_order:
                if table_name not in wb.sheetnames:
                    continue
                ws = wb[table_name]
                headers = [cell.value for cell in ws[1]]
                if not headers:
                    continue
                if not all(isinstance(h, str) and re.match(r"^\w+$", h) for h in headers):
                    continue

                cols = ", ".join(headers)
                placeholders = ", ".join(["?"] * len(headers))

                for row_idx in range(2, ws.max_row + 1):
                    values = [cell.value for cell in ws[row_idx]]
                    conn.execute(
                        f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})",
                        values,
                    )
                    total_inserted += 1

            conn.commit()

        clear_schema_cache()
        clear_prompt_cache()

        return {"ok": True, "total_inserted": total_inserted}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.post("/reset-database")
async def reset_database():
    try:
        reset_db(SCHEMA_PATH)
        seed_initial_data()
        clear_schema_cache()
        clear_prompt_cache()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# --- User Management ---

@router.get("/users")
async def list_users():
    users = _query(
        "SELECT id, first_name, last_name, email, username, user_role, role_title, status "
        "FROM people WHERE is_user = 1 ORDER BY id"
    )
    return {"users": users}


class AddUserRequest(BaseModel):
    first_name: str
    last_name: str
    email: str
    role: str = "user"
    requesting_person_id: int


@router.post("/users")
async def add_user(req: AddUserRequest):
    if req.role not in ("user", "admin"):
        return {"error": "Role must be 'user' or 'admin'"}

    requester = _query(
        "SELECT user_role FROM people WHERE id = ? AND is_user = 1", [req.requesting_person_id]
    )
    if not requester:
        return {"error": "Requesting user not found"}
    requester_role = requester[0]["user_role"]

    if req.role == "admin" and requester_role != "owner":
        return {"error": "Only the owner can add admins"}
    if requester_role not in ("owner", "admin"):
        return {"error": "Only owner or admin can add users"}

    base_username = req.first_name.lower().strip()
    username = base_username
    suffix = 1
    while _query("SELECT 1 FROM people WHERE username = ?", [username]):
        suffix += 1
        username = f"{base_username}{suffix}"

    conn = get_connection()
    with _lock:
        conn.execute(
            "INSERT INTO people (first_name, last_name, email, is_user, username, user_role, status) "
            "VALUES (?, ?, ?, 1, ?, ?, 'active')",
            [req.first_name.strip(), req.last_name.strip(), req.email.strip(), username, req.role],
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()

    email_result = send_email(
        to_email=req.email.strip(),
        subject=f"You've been added to TrueCore.cloud as {req.role}",
        body=(
            f"Hi {req.first_name.strip()},\n\n"
            f"You've been added to TrueCore.cloud as a {req.role}.\n"
            f"Your username is: {username}\n\n"
            f"Access TrueCore.cloud here: {APP_URL}\n\n"
            f"— TrueCore.cloud"
        ),
        to_name=f"{req.first_name.strip()} {req.last_name.strip()}",
    )

    return {"ok": True, "person_id": new_id, "username": username, "email_sent": email_result}


@router.delete("/users/{person_id}")
async def remove_user(person_id: int, requesting_person_id: int = Query(...)):
    target = _query("SELECT user_role FROM people WHERE id = ? AND is_user = 1", [person_id])
    if not target:
        return {"error": "User not found"}
    if target[0]["user_role"] == "owner":
        return {"error": "Cannot remove the owner"}

    requester = _query(
        "SELECT user_role FROM people WHERE id = ? AND is_user = 1", [requesting_person_id]
    )
    if not requester:
        return {"error": "Requesting user not found"}
    requester_role = requester[0]["user_role"]
    target_role = target[0]["user_role"]

    if requester_role == "admin" and target_role == "admin":
        return {"error": "Admins cannot remove other admins"}
    if requester_role not in ("owner", "admin"):
        return {"error": "Insufficient permissions"}

    conn = get_connection()
    with _lock:
        conn.execute(
            "UPDATE people SET is_user = 0, status = 'inactive' WHERE id = ?", [person_id]
        )
        conn.commit()

    return {"ok": True}


class ChangeRoleRequest(BaseModel):
    new_role: str
    requesting_person_id: int


@router.patch("/users/{person_id}/role")
async def change_user_role(person_id: int, req: ChangeRoleRequest):
    if req.new_role not in ("user", "admin"):
        return {"error": "Role must be 'user' or 'admin'"}

    requester = _query(
        "SELECT user_role FROM people WHERE id = ? AND is_user = 1", [req.requesting_person_id]
    )
    if not requester or requester[0]["user_role"] != "owner":
        return {"error": "Only the owner can change roles"}

    target = _query("SELECT user_role FROM people WHERE id = ? AND is_user = 1", [person_id])
    if not target:
        return {"error": "User not found"}
    if target[0]["user_role"] == "owner":
        return {"error": "Cannot change the owner's role"}

    conn = get_connection()
    with _lock:
        conn.execute("UPDATE people SET user_role = ? WHERE id = ?", [req.new_role, person_id])
        conn.commit()

    return {"ok": True}
